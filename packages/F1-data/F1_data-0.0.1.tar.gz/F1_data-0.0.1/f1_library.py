
import fastf1
import fastf1.plotting
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import pandas as pd


def data_arrange(driver,col,i):
	sel=session.laps.pick_driver(driver)
	laps=list(sel['LapTime'])
	comp=list(sel['Compound'])


	k=0
	i=2
	for element in laps:
		cell=col[j]+str(i)
		ws[cell]=element



		if comp[k]=='SOFT':
			ws[cell].fill = PatternFill(start_color="eb0e2b", end_color="eb0e2b", fill_type = "solid")
		elif comp[k]=='MEDIUM':
			ws[cell].fill = PatternFill(start_color="ebc60e", end_color="ebc60e", fill_type = "solid")
		elif comp[k]=='INTERMEDIATE':
			ws[cell].fill = PatternFill(start_color="4287f5", end_color="4287f5", fill_type = "solid")
		elif comp[k]=='WET':
			ws[cell].fill = PatternFill(start_color="35b033", end_color="35b033", fill_type = "solid")

		i+=1
		k+=1




def session_all_laps(year,gp,session):

	wb=Workbook()
	ws=wb.active


	i=2
	j=1

	cell=ws.cell(i,j)

	while i < 100 and j<100:
		cell.number_format= 'mm:ss,000'
		if i <100:
			i+=1
		else:
			j+=1
			i=2

	column=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','Z']

	session = fastf1.get_session(year,gp,s)

	session.load()

	driver_list=list(pd.unique(session.laps['Driver']))

	ws.append(driver_list)

	j=0

	for element in driver_list:
		data_arrange(element,column,j)

		j+=1

