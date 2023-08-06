from openpyxl.comments import Comment
from openpyxl.styles import Protection,Font,Alignment,numbers
from openpyxl.styles.colors import Color
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fills import PatternFill

style={
    'description_column_width':70,
    'value_column_width':70,
    'display_title':2,
    'data_start_row':4,
    "COLUMN_WIDTH":{
        "date":15,
        'name':20,
        'fullname':30,
        'country':20,
        'relation':15,
        'address':30,
        'type':20,
        'email':25,
        'title':20,
        'city':15,
        'province':15,
        'status':20,
        'field_of_study':20,
        'duties':50,
        'brief':30    
    },
    "default_format":{
        "fill":PatternFill(
            fill_type=None
            ),
        "font":Font(name="Arial (Body)",size=11,color="00000000"),
        "alignment":Alignment(  
                                horizontal='center',
                                vertical='center',
                                text_rotation=0,
                                wrap_text=True,
                                shrink_to_fit=False,
                                indent=0
                            ),
        "border":Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                    ),
        'number_format':numbers.FORMAT_TEXT,
        'protection':Protection(locked=True)
    },
    "sheet_title_format":{
        "fill":PatternFill(
            fill_type='solid',
            fgColor='f4f3ee'
            ),
        "font":Font(name="Arial (Headings)",size=24,bold=True,color="272727"),
        "alignment":Alignment(  
                                horizontal='center',
                                vertical='center'
                            )
    },
    "column_titlet_format":{
        "fill":PatternFill(
            fill_type='solid',
            fgColor='f4f3ee'
            ),
        "font":Font(size=14,bold=True,color="00000000"),
        "alignment":Alignment(  
                                horizontal='center',
                                vertical='center'
                            )
    },    
    "description_format":{
        "fill":PatternFill(
            fill_type='solid',
            fgColor='f4f3ee'
            ),
        "font":Font(size=14),
        "alignment":Alignment(  
                                horizontal='right',
                                vertical='center'
                            )
    },
    "value_format":{
        "fill":PatternFill(
            fill_type='solid',
            fgColor='ffffffff'
            ),
        "font":Font(size=14,color="719c75"),
        "alignment":Alignment(  
                                horizontal='left',
                                vertical='center',
                                wrap_text=True
                            ),
        'protection':Protection(locked=False)
    },
    "table_format":{
        "fill":Color(rgb="AA000000")
    },
    'date_format':"YYYY-MM-DD",
    "GUIDE":{
        "title":"Guide",
        "contents":[
            "This excel includes 2 sheets (contact, table-job_post) in addtion to this one. ",
            "For information input, if you don't know how to enter, you can move your mouse to the value field, and the help will pop up",
            "Please be noticed that date format must be like 2021-12-31",
            "if some value fields need range, please input two number seperated by comma. Exp: 26, 29",
            "If some value fields need a list of sentence, you can edit in other editor, and then copy it to the field. "
        ]
    }
}

hidden_sheets=['admin','system']