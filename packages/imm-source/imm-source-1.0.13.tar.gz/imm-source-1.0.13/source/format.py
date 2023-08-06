from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.views import BookView
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import Protection
from openpyxl.workbook.protection import WorkbookProtection
import argparse,os
from copy import copy
from source.style import style

class Format(object):
    
    def __init__(self,workbook)->None:
        self.wb=workbook
        self.wb.views=[BookView(xWindow=0, yWindow=0, windowWidth=18140, windowHeight=15540)]
        self.sheetNames=self.wb.sheetnames
        self.sheets=[sheetName for sheetName in self.sheetNames if sheetName.lower().startswith("info-")]
        self.tables=[sheetName for sheetName in self.sheetNames if sheetName.lower().startswith("table-")]
        self.description_column_width=style.get('description_column_width',70)
        self.value_column_width=style.get('value_column_width',70)
        self.data_start_row=style.get('data_start_row',4)
        self.variable_start_row=style.get('variable_start_row',3)
        self.display_row=style.get('display_title',2)  #get display row number from style
        
    def setCellStyle(self,cell,style,default_style):
        cell.font = copy( {**default_style,**style}).get("font")
        cell.border = copy( {**default_style,**style}).get("border")
        cell.fill = copy( {**default_style,**style}).get("fill")
        cell.alignment = copy( {**default_style,**style}).get("alignment")
        cell.number_format=copy( {**default_style,**style}).get("number_format")
        cell.protection = copy(style.get('protection'))
        
    
    # This works for both info and table
    def setSheetTitle(self,sheet):
        self.setCellStyle(sheet['A1'],style.get('sheet_title_format'),style.get('default_format'))
    
    # This works for both info and table
    def setColumnTitles(self,sheet):
        max_col=sheet.max_column
        for row in range(self.display_row,4): 
            for col in range(1,max_col+1):
                cell=sheet[get_column_letter(col)+str(row)]
                self.setCellStyle(cell,style.get('column_titlet_format'),style.get('default_format'))
            # hide variable row
        sheet.row_dimensions[self.variable_start_row].hidden = True # hide the variable row
                
    def setTableDataAreas(self,sheet):
        max_col=sheet.max_column
        max_row=sheet.max_row
        for row in range(self.data_start_row,max_row+1):
            for col in range(1,max_col+1):
                cell=sheet.cell(row,col)
                self.setCellStyle(cell,style.get('value_format'),style.get('default_format'))
                # find row 3's variable title, if title includes date, then assign it number format
                variable_title=sheet.cell(self.variable_start_row,col)
                if variable_title.value and 'date' in variable_title.value.lower():
                    cell.number_format=style.get("date_format")

    # get column cells in info-sheet by proving column title
    def _getColumnCells(self,sheet,column_title):
        variable_titles=list(sheet.values)[2] # row 3 is variable row
        try:
            col=variable_titles.index(column_title)+1
            cells=[]
            for row in range(self.data_start_row,sheet.max_row+1):
                cells.append(sheet.cell(row,col))
            return cells
        except:
            raise ValueError(f'{column_title} is not existed')
    
    # for info table, we may have different style for 'description' column and 'value' column
    def setDescriptColumn(self,sheet):
        for cell in self._getColumnCells(sheet,'description'):
            self.setCellStyle(cell,style.get('description_format'),style.get('default_format'))
            sheet.column_dimensions[cell.column_letter].width = self.description_column_width #set description column width
    
    # for info table value column
    def setSheetValueColumn(self,sheet):
        for i,cell in enumerate(self._getColumnCells(sheet,'value')):
            self.setCellStyle(cell,style.get('value_format'),style.get('default_format'))
            sheet.column_dimensions[cell.column_letter].width = self.value_column_width # set value column width
            # if 
            variable_cell=self._getColumnCells(sheet,'variable')[i] # get same row's variable cell
            if 'date' in variable_cell.value.lower():
                cell.number_format=style.get("date_format")
                
    #get column width from style
    def getColumnWideth(self,column_name):
        for key in style.get('COLUMN_WIDTH').keys():
            if column_name and key in column_name:
                return style.get('COLUMN_WIDTH').get(key)
        return 10 # if no definition, return default as 10
    
    # set table column width            
    def setTableColumnWidth(self,table):
        for col in table.columns:
            display_cell=col[self.display_row]
            col_width=self.getColumnWideth(display_cell.value)    #col[2] is cell of variable. 
            table.column_dimensions[display_cell.column_letter].width = col_width

    #hide variable, and tag columns
    def hideColumn(self,sheet,columns=['variable','tag']):
        for col_name in columns:
            for cell in self._getColumnCells(sheet,col_name):
                sheet.column_dimensions[cell.column_letter].hidden = True
    #protect 
    def protectSheets(self,password):
        self.wb.security=WorkbookProtection(password, lockStructure = True)
        for sheet_name in self.wb.sheetnames:
            workSheet=self.wb[sheet_name]
            workSheet.protection.sheet=True
            workSheet.protection.password=password
            workSheet.protection.formatColumns=False
            workSheet.protection.formatRows=False
            workSheet.protection.enable()
        
        # unprotect table sheets from row 4
        for table_name in self.tables:
            table=self.wb[table_name]
            for row in range(self.data_start_row,105):
                for col in range(1,table.max_column+1):
                    table.cell(row,col).protection = Protection(locked=False)
        
    
    # fromat all 
    def setAll(self):
        for sheet in self.sheets:
            self.setSheetTitle(self.wb[sheet])
            self.setColumnTitles(self.wb[sheet])
            self.setDescriptColumn(self.wb[sheet])
            self.setSheetValueColumn(self.wb[sheet])
            self.hideColumn(self.wb[sheet])
            
        for table in self.tables:
            self.setSheetTitle(self.wb[table])
            self.setColumnTitles(self.wb[table])
            self.setTableDataAreas(self.wb[table])
            self.setTableColumnWidth(self.wb[table])
    # protect all
    def protectAll(self):    
        password=input("please input password:")
        self.protectSheets(password)